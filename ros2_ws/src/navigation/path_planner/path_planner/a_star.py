#
# MOBILE ROBOTS - FI-UNAM, 2026-2
# PATH PLANNING BY A-STAR
#
# Instructions:
# Write the code necessary to plan a path using an
# occupancy grid and the A* algorithm
# MODIFY ONLY THE SECTIONS MARKED WITH THE 'TODO' COMMENT
#
import rclpy
from rclpy.node import Node
from rclpy.time import Time, Duration
from geometry_msgs.msg import PoseStamped, Pose, Point
from nav_msgs.msg import Path
from nav_msgs.srv import *
from builtin_interfaces.msg import Duration
from collections import deque
import numpy
import heapq
import math
# ── NUEVO ────────────────────────────────────────────────────
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
EXCEL_FILE = "resultados_astar.xlsx"
# ────────────────────────────────────────────────────────────

NAME = "DOMÍNGUEZ PALACIOS JESÚS ALEJANDRO"

class AStarNode(Node):
    def a_star(self, start_r, start_c, goal_r, goal_c, grid_map, cost_map, use_diagonals):
        [height, width] = grid_map.shape
        in_open_list   = numpy.full(grid_map.shape, False)
        in_closed_list = numpy.full(grid_map.shape, False)
        g_values       = numpy.full(grid_map.shape, float("inf"))
        f_values       = numpy.full(grid_map.shape, float("inf"))
        parent_nodes   = numpy.full((grid_map.shape[0],grid_map.shape[1],2),-1)
        open_list = []
        if use_diagonals: #Every adjacent node has: [row_offset, col_offset, cost]
            adjacents = [[1,0,1],[0,1,1],[-1,0,1],[0,-1,1], [1,1,1.414], [-1,1,1.414], [-1,-1,1.414],[1,-1,1.414]]
        else:
            adjacents = [[1,0,1],[0,1,1],[-1,0,1],[0,-1,1]]

        heapq.heappush(open_list, (0, [start_r, start_c]))
        in_open_list[start_r, start_c] = True
        g_values    [start_r, start_c] = 0
        [row, col]= [start_r, start_c]   #Current node
        #
        # TODO:
        # Implement the A* algorithm for path planning
        # Map is considered to be a 2D array and start and goal positions
        # are given as row-col pairs. You can follow these steps:
        #
        # WHILE open list is not empty and current is different from goal:
        #     Get current node [row,col] from open list (see heapq.heappop function)
        #     Mark current node as 'in_closed_list'
        #     For [r,c,cost] in adjacent nodes:
        #         Get r,c indices of neighbours of current node (check content of adjacents)
        #         Discard if r,c is out of map, occupied, unknonw or in closed list, and continue
        #         get a g-value g as: g-value of current node + dist + cost of neighbour r,c
        #         Calculate heuristic 
        #         Calculate f-value
        #         IF g < g_value of neighbour r,c:
        #             set g as g_value of neighbour r,c
        #             set f as f_value of neighbour r,c
        #             SET current node row,col as parent of neighbour r,c
        #         If neighbour r,c is not in open list:
        #             mark r,c as 'in_open_list'
        #             add r,c to open list (check heapq.heappush)
        #
        while len(open_list) > 0 and [row, col] != [goal_r, goal_c]:
            # Get current node with lowest f-value from open list
            _, [row, col] = heapq.heappop(open_list)
            # Mark current node as visited
            in_closed_list[row, col] = True
            # Explore each neighbour
            for [dr, dc, move_cost] in adjacents:
                r = row + dr
                c = col + dc
                # Discard if out of map bounds
                if r < 0 or r >= height or c < 0 or c >= width:
                    continue
                # Discard if occupied (100), unknown (-1), or already in closed list
                if grid_map[r, c] != 0 or in_closed_list[r, c]:
                    continue
                # g = g_current + movement_cost + cell_cost (normalised 0-1)
                g = g_values[row, col] + move_cost + cost_map[r, c] / 100.0
                # Heuristic: Euclidean distance to goal
                h = math.sqrt((goal_r - r)**2 + (goal_c - c)**2)
                # f = g + h
                f = g + h
                # Update if a better path to this neighbour is found
                if g < g_values[r, c]:
                    g_values[r, c]     = g
                    f_values[r, c]     = f
                    parent_nodes[r, c] = [row, col]
                # Add to open list if not already there
                if not in_open_list[r, c]:
                    in_open_list[r, c] = True
                    heapq.heappush(open_list, (f, [r, c]))
        #
        # END OF WHILE
        #
        path = []
        while parent_nodes[goal_r, goal_c][0] != -1:
            path.insert(0, [goal_r, goal_c])
            [goal_r, goal_c] = parent_nodes[goal_r, goal_c]
        return path

    # ── NUEVO: guarda cada experimento en Excel ──────────────
    def save_to_excel(self, sx, sy, gx, gy, diagonals, delta_ms, n_points, success):
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        except:
            # Primera vez: crear archivo con cabeceras
            wb = Workbook()
            ws = wb.active
            ws.title = "Astar Experimentos"
            BLUE = PatternFill("solid", start_color="1F4E79")
            thin = Side(style="thin", color="BFBFBF")
            BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
            CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
            LEFT = Alignment(horizontal="left",   vertical="center")
            # Titulo
            ws.merge_cells("A1:J1")
            c           = ws["A1"]
            c.value     = "EXPERIMENTOS - Algoritmo A-Star  |  Robotica Movil FI-UNAM 2026-2"
            c.font      = Font(name="Arial", bold=True, size=12, color="1F4E79")
            c.alignment = LEFT
            # Cabeceras
            headers = ["#", "Inicio (x,y)", "Meta (x,y)", "Diagonales",
                       "inflation_r", "cost_r", "Tiempo (ms)", "N Puntos",
                       "Exito", "Resultado"]
            widths  = [5, 14, 14, 11, 11, 9, 13, 10, 8, 14]
            for col, (h, w) in enumerate(zip(headers, widths), 1):
                ws.column_dimensions[get_column_letter(col)].width = w
                cell           = ws.cell(row=3, column=col, value=h)
                cell.fill      = BLUE
                cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=9)
                cell.alignment = CTR
                cell.border    = BORD
            ws.row_dimensions[3].height = 28
            ws.freeze_panes = "A4"

        # Leer parametros declarados en el nodo
        inf_r  = self.get_parameter('inflation_radius').get_parameter_value().double_value
        cost_r = self.get_parameter('cost_radius').get_parameter_value().double_value

        exp_num = ws.max_row - 2
        row     = ws.max_row + 1
        ALT_B   = PatternFill("solid", start_color="D6E4F0")
        thin    = Side(style="thin", color="BFBFBF")
        BORD    = Border(left=thin, right=thin, top=thin, bottom=thin)
        fill    = ALT_B if exp_num % 2 == 0 else None

        vals = [
            exp_num,
            f"({sx:.2f}, {sy:.2f})",
            f"({gx:.2f}, {gy:.2f})",
            "Si" if diagonals else "No",
            inf_r,
            cost_r,
            round(delta_ms, 2),
            n_points,
            1 if success else 0,
            "Exito" if success else "Sin ruta"
        ]
        for col, v in enumerate(vals, 1):
            cell           = ws.cell(row=row, column=col, value=v)
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = BORD
            if fill: cell.fill = fill

        wb.save(EXCEL_FILE)
    # ── FIN save_to_excel ────────────────────────────────────

    def get_maps(self):
        self.get_logger().info("Waiting for inflated map service...")
        while not self.clt_inflated_map.wait_for_service(timeout_sec=1.0):
            self.get_logger().info('Waiting for inflated map service...')
        self.get_logger().info("Inflated map service is now available...")
        self.get_logger().info("Waiting for cost map service...")
        while not self.clt_cost_map.wait_for_service(timeout_sec=1.0):
            self.get_logger().info('Waiting for cost map service...')
        self.get_logger().info("Cost map service is now available...")

        self.get_logger().info("Trying to get inflated map...")
        future = self.clt_inflated_map.call_async(GetMap.Request())
        rclpy.spin_until_future_complete(self, future)
        inflated_map = future.result().map
        self.get_logger().info("Got inflated map.")
        self.get_logger().info("Trying to get cost map...")
        future = self.clt_cost_map.call_async(GetMap.Request())
        rclpy.spin_until_future_complete(self, future)
        cost_map= future.result().map
        self.get_logger().info("Got cost map.")
        return [inflated_map, cost_map]

    def get_path_msg(self, path, res, zx, zy):
        msg_path = Path()
        msg_path.header.frame_id = "map"
        msg_path.header.stamp = self.get_clock().now().to_msg()
        msg_path.poses = []
        for [r,c] in path:
            msg_path.poses.append(PoseStamped(pose=Pose(position=Point(x=(c*res + zx), y=(r*res + zy)))))
        return msg_path

    def callback_a_star(self, req, resp):
        info = self.inflated_map.info
        res = info.resolution
        [sx, sy] = [req.start.pose.position.x, req.start.pose.position.y]
        [gx, gy] = [req.goal .pose.position.x, req.goal .pose.position.y]
        [zx, zy] = [self.inflated_map.info.origin.position.x, self.inflated_map.info.origin.position.y]
        use_diagonals = self.get_parameter('diagonals').get_parameter_value().bool_value
        inflated_grid = numpy.reshape(numpy.asarray(self.inflated_map.data), (info.height, info.width))
        cost_grid     = numpy.reshape(numpy.asarray(self.cost_map.data)    , (info.height, info.width))
        
        self.get_logger().info("Planning path by A* from " + str([sx, sy])+" to "+str([gx, gy]))
        start_time = self.get_clock().now()
        path = self.a_star(int((sy-zy)/res), int((sx-zx)/res), int((gy-zy)/res), int((gx-zx)/res),
                           inflated_grid, cost_grid, use_diagonals)
        end_time = self.get_clock().now()
        delta_ms = (end_time.nanoseconds - start_time.nanoseconds)/1e6
        if len(path) > 0:
            self.get_logger().info("Path planned after " + str(delta_ms) + " ms with " +  str(len(path)) + " points")
        else:
            self.get_logger().info("Cannot plan path from  " + str([sx, sy])+" to "+str([gx, gy]) + " :'(")

        # ── NUEVO: guardar resultado en Excel ────────────────
        self.save_to_excel(sx, sy, gx, gy, use_diagonals, delta_ms, len(path), len(path) > 0)
        self.get_logger().info(f"Resultado guardado en {EXCEL_FILE}")
        # ────────────────────────────────────────────────────

        self.msg_path = self.get_path_msg(path, res, zx, zy)
        resp.plan = self.msg_path
        return resp

    def callback_timer(self):
        self.pub_path.publish(self.msg_path)
            
    def __init__(self):
        super().__init__("a_star_node")
        self.get_logger().info("INITIALIZING A STAR NODE - " + NAME)
        self.clt_inflated_map = self.create_client(GetMap, '/get_inflated_map')
        self.clt_cost_map     = self.create_client(GetMap, '/get_cost_map')
        
        [self.inflated_map, self.cost_map] = self.get_maps()
        self.declare_parameter('diagonals', False)
        self.declare_parameter('inflation_radius', 0.2)  # ── NUEVO
        self.declare_parameter('cost_radius', 0.1)        # ── NUEVO
        self.srv_plan_path = self.create_service(GetPlan, '/path_planning/plan_path', self.callback_a_star)
        self.pub_path = self.create_publisher(Path, '/path_planning/path', 10)
        self.msg_path = Path()
        self.timer = self.create_timer(0.5, self.callback_timer)
            
def main(args=None):
    rclpy.init(args=args)
    a_star_node = AStarNode()
    rclpy.spin(a_star_node)
    a_star_node.destroy_node()
    rclpy.shutdown()

    
if __name__ == '__main__':
    main()
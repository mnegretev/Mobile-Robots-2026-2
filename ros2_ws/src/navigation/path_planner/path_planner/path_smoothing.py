#
# MOBILE ROBOTS - FI-UNAM, 2026-2
# PATH SMOOTHING BY GRADIENT DESCEND
#
# Instructions:
# Write the code necessary to smooth a path using the gradient descend algorithm.
# MODIFY ONLY THE SECTIONS MARKED WITH THE 'TODO' COMMENT
#
import rclpy
from rclpy.node import Node
from nav_msgs.msg import Path
from geometry_msgs.msg import Pose, PoseStamped, Point
from navig_msgs.srv import ProcessPath
import numpy
import openpyxl
import os
from datetime import datetime

NAME = "DOMÍNUGEZ PALACIOS JESÚS ALEJANDRO"

class PathSmoothingNode(Node):

    def smooth_path(self, Q, w1, w2, max_steps):
        P       = numpy.copy(Q)
        tol     = 0.00001
        nabla   = numpy.full(Q.shape, float("inf"))
        epsilon = 0.1
        steps   = 0

        # Verificar que la ruta tenga al menos 3 puntos para poder suavizar
        if len(P) < 3:
            self.get_logger().warn("La ruta tiene menos de 3 puntos, no se puede suavizar.")
            return P

        # Los puntos extremos (inicio y fin) se fijan en cero
        # para que no se muevan durante el suavizado
        nabla[0]  = 0
        nabla[-1] = 0

        # Repetir mientras el gradiente sea grande O no se alcance el límite de pasos
        while numpy.linalg.norm(nabla) > tol and steps < max_steps:

            # Calcular el gradiente para cada punto interior i ∈ [1, n-2]
            for i in range(1, len(P) - 1):
                nabla[i] = w1 * (2 * P[i] - P[i - 1] - P[i + 1]) \
                         + w2 * (P[i] - Q[i])
                # w1: penaliza la curvatura (suavidad de la ruta)
                # w2: penaliza alejarse de la ruta original (fidelidad)

            # Paso de descenso de gradiente (solo puntos interiores)
            P[1:-1] -= epsilon * nabla[1:-1]

            steps += 1

        return P

    def export_to_excel(self, Q, P, w1, w2, steps, delta_ms):
        filename = os.path.expanduser("~/path_smoothing_resultados.xlsx")

        # Si el archivo ya existe, abrirlo para agregar datos; si no, crearlo
        if os.path.exists(filename):
            wb = openpyxl.load_workbook(filename)
        else:
            wb = openpyxl.Workbook()

        # --- Hoja 1: Resumen de parámetros por ejecución ---
        if "Resumen" not in wb.sheetnames:
            ws_res = wb.create_sheet("Resumen", 0)
            ws_res.append(["Fecha", "w1", "w2", "Pasos", "Puntos ruta", "Tiempo (ms)"])
        else:
            ws_res = wb["Resumen"]

        # Agregar una fila resumen por cada ejecución
        ws_res.append([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            w1, w2, steps, len(Q), round(delta_ms, 3)
        ])

        # --- Hoja 2: Solo inicio, fin y algunos puntos clave de la ruta ---
        sheet_name = f"w1={w1}_w2={w2}"
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws_path = wb.create_sheet(sheet_name)
        ws_path.append(["Punto", "X orig", "Y orig", "X suav", "Y suav"])

        # Guardar solo inicio, fin y cada 10 puntos intermedios
        indices = list(range(0, len(Q), max(1, len(Q)//10))) + [len(Q)-1]
        indices = sorted(set(indices))
        for i in indices:
            ws_path.append([i,
                round(Q[i,0],4), round(Q[i,1],4),
                round(P[i,0],4), round(P[i,1],4)])

        # Eliminar la hoja vacía por defecto si existe
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        wb.save(filename)
        self.get_logger().info("Excel actualizado en: " + filename)

    def callback_smooth_path(self, request, response):
        w1    = self.get_parameter('w1').get_parameter_value().double_value
        w2    = self.get_parameter('w2').get_parameter_value().double_value
        steps = self.get_parameter('steps').get_parameter_value().integer_value

        self.get_logger().info("Smoothing path with params: " + str([w1, w2, steps]))

        start_time = self.get_clock().now()
        Q = numpy.asarray([[p.pose.position.x, p.pose.position.y] for p in request.path.poses])
        P = self.smooth_path(Q, w1, w2, steps)
        end_time = self.get_clock().now()

        delta_ms = (end_time.nanoseconds - start_time.nanoseconds)/1e6
        self.get_logger().info("Path smoothed after " + str(delta_ms) + " ms")

        # Exportar los resultados a Excel
        self.export_to_excel(Q, P, w1, w2, steps, delta_ms)

        self.msg_smooth_path.header.frame_id = request.path.header.frame_id
        self.msg_smooth_path.header.stamp    = self.get_clock().now().to_msg()
        self.msg_smooth_path.poses = []

        for i in range(len(request.path.poses)):
            p = PoseStamped()
            p.pose.position.x = P[i,0]
            p.pose.position.y = P[i,1]
            self.msg_smooth_path.poses.append(p)

        self.pub_smooth_path.publish(self.msg_smooth_path)
        response.processed_path = self.msg_smooth_path
        return response

    def __init__(self):
        super().__init__("path_smoothing_node")
        self.get_logger().info("INITIALIZING PATH SMOOTHING NODE - " + NAME)
        self.declare_parameter('w1',    0.9)
        self.declare_parameter('w2',    0.1)
        self.declare_parameter('steps', 10000)
        self.srv_smooth_path = self.create_service(ProcessPath, '/path_planning/smooth_path', self.callback_smooth_path)
        self.pub_smooth_path = self.create_publisher(Path, '/path_planning/smoothed_path', 10)
        self.msg_smooth_path = Path()

def main(args=None):
    rclpy.init(args=args)
    path_smoothing_node = PathSmoothingNode()
    rclpy.spin(path_smoothing_node)
    path_smoothing_node.destroy_node()
    rclpy.shutdown()

if __name__ == '__main__':
    main()